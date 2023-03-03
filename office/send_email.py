import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage

import openpyxl


def send_email(excel_path):
    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 获取电子邮件地址、主题、邮件主体、附件路径、发送人和祝福语
    sTo = ws.cell(row=2, column=1).value
    sSubject = "主题图片"
    sX = ws.cell(row=2, column=2).value
    sBody = sX + "您好，您要的图片在附件里"
    sGreeting = ws.cell(row=2, column=4).value
    sPath = ws.cell(row=2, column=5).value

    # 创建电子邮件对象
    msg = MIMEMultipart()
    msg['To'] = sTo
    msg['Subject'] = sSubject

    # 添加文本内容
    body = sBody + '\n\n' + sGreeting
    msg.attach(MIMEText(body, 'plain'))

    # 添加图片附件
    with open(sPath, 'rb') as f:
        img_data = f.read()
    image = MIMEImage(img_data, name=sPath)
    msg.attach(image)

    # 发送电子邮件
    try:
        server = smtplib.SMTP('localhost')
        server.sendmail('', sTo, msg.as_string())
        print("电子邮件发送成功")
    except:
        print("Error: 无法发送电子邮件")

    # 关闭SMTP对象
    server.quit()


def send_email_prompt():
    # 提示用户输入Excel文件路径
    excel_path = input("请输入Excel文件路径: ")

    # 调用send_email()函数发送电子邮件
    send_email(excel_path)


send_email_prompt()
